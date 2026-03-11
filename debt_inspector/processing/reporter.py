"""
Генерация отчётов: JSON, Excel, консольный вывод.
"""

import json
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rich.console import Console
from rich.table import Table
from rich.panel import Panel

from debt_inspector.models.report import DebtReport

console = Console()

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_json(report: DebtReport, path: str | Path) -> Path:
    """Сохраняет отчёт в JSON."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    data = {
        "summary": report.summary(),
        "enforcements": [e.model_dump(mode="json") for e in report.enforcements],
        "court_cases": [c.model_dump(mode="json") for c in report.court_cases],
        "bankruptcies": [b.model_dump(mode="json") for b in report.bankruptcies],
    }

    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return path


def export_excel(report: DebtReport, path: str | Path) -> Path:
    """Сохраняет отчёт в Excel."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Лист: Сводка ---
    ws = wb.active
    ws.title = "Сводка"
    summary = report.summary()
    ws.append(["Отчёт по проверке должника"])
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for key, val in summary.items():
        ws.append([_ru_label(key), str(val)])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    # --- Лист: Исполнительные производства ---
    if report.enforcements:
        ws_ip = wb.create_sheet("ИП (ФССП)")
        headers = ["Номер", "Дата", "Предмет", "Сумма", "Статус", "Отдел", "Пристав", "Взыскатель"]
        ws_ip.append(headers)
        _style_header(ws_ip, len(headers))

        for e in report.enforcements:
            ws_ip.append([
                e.number,
                str(e.date_opened) if e.date_opened else "",
                e.subject,
                e.amount,
                e.status.value if e.status else "",
                e.department,
                e.bailiff,
                e.claimant,
            ])
        _auto_width(ws_ip)

    # --- Лист: Судебные дела ---
    if report.court_cases:
        ws_court = wb.create_sheet("Судебные дела")
        headers = ["Номер дела", "Суд", "Дата подачи", "Предмет", "Сумма", "Статус", "Истец", "Ответчик"]
        ws_court.append(headers)
        _style_header(ws_court, len(headers))

        for c in report.court_cases:
            ws_court.append([
                c.case_number,
                c.court_name,
                str(c.date_filed) if c.date_filed else "",
                c.subject,
                c.amount,
                c.status.value if c.status else "",
                c.plaintiff,
                c.defendant,
            ])
        _auto_width(ws_court)

    # --- Лист: Банкротства ---
    if report.bankruptcies:
        ws_bank = wb.create_sheet("Банкротства")
        headers = ["Должник", "ИНН", "Дело", "Фаза", "Арб. управляющий", "Суд", "Дата решения"]
        ws_bank.append(headers)
        _style_header(ws_bank, len(headers))

        for b in report.bankruptcies:
            ws_bank.append([
                b.debtor_name,
                b.debtor_inn,
                b.case_number,
                b.phase.value if b.phase else "",
                b.arbitration_manager,
                b.court_name,
                str(b.date_decision) if b.date_decision else "",
            ])
        _auto_width(ws_bank)

    wb.save(path)
    return path


def print_report(report: DebtReport):
    """Красивый вывод в консоль."""
    console.print()
    console.print(
        Panel(
            f"[bold]{report.search_params.display_name}[/bold]\n"
            f"Проверка: {report.checked_at.strftime('%d.%m.%Y %H:%M')}",
            title="Отчёт по должнику",
            border_style="blue",
        )
    )

    # Сводка
    summary_table = Table(show_header=False, box=None, padding=(0, 2))
    summary_table.add_column(style="cyan")
    summary_table.add_column(style="bold")
    summary_table.add_row("Исполнительных производств:", str(len(report.enforcements)))
    summary_table.add_row("Долг по ИП:", f"{report.total_enforcement_debt:,.2f} руб.")
    summary_table.add_row("Судебных дел:", str(len(report.court_cases)))
    summary_table.add_row("Сумма исков:", f"{report.total_court_claims:,.2f} руб.")
    summary_table.add_row("Банкротств:", str(len(report.bankruptcies)))
    summary_table.add_row(
        "Активное банкротство:",
        "[red]ДА[/red]" if report.has_active_bankruptcy else "[green]НЕТ[/green]",
    )
    console.print(summary_table)

    # ИП
    if report.enforcements:
        console.print()
        table = Table(title="Исполнительные производства (ФССП)", show_lines=True)
        table.add_column("Номер", style="cyan", max_width=25)
        table.add_column("Дата")
        table.add_column("Сумма", justify="right", style="red")
        table.add_column("Статус")
        table.add_column("Предмет", max_width=40)

        for e in report.enforcements:
            status_style = {"active": "red", "finished": "green", "suspended": "yellow"}.get(
                e.status.value, ""
            )
            table.add_row(
                e.number or "—",
                str(e.date_opened) if e.date_opened else "—",
                f"{e.amount:,.2f}" if e.amount else "—",
                f"[{status_style}]{e.status.value}[/{status_style}]",
                (e.subject or "—")[:40],
            )
        console.print(table)

    # Суды
    if report.court_cases:
        console.print()
        table = Table(title="Судебные дела (КАД Арбитр)", show_lines=True)
        table.add_column("Дело", style="cyan")
        table.add_column("Суд", max_width=30)
        table.add_column("Дата")
        table.add_column("Сумма", justify="right", style="red")
        table.add_column("Статус")

        for c in report.court_cases:
            table.add_row(
                c.case_number or "—",
                (c.court_name or "—")[:30],
                str(c.date_filed) if c.date_filed else "—",
                f"{c.amount:,.2f}" if c.amount else "—",
                c.status.value,
            )
        console.print(table)

    # Банкротства
    if report.bankruptcies:
        console.print()
        table = Table(title="Банкротства (ЕФРСБ)", show_lines=True)
        table.add_column("Должник", style="cyan", max_width=30)
        table.add_column("ИНН")
        table.add_column("Дело")
        table.add_column("Фаза", style="yellow")
        table.add_column("Арб. управляющий", max_width=25)

        for b in report.bankruptcies:
            table.add_row(
                (b.debtor_name or "—")[:30],
                b.debtor_inn or "—",
                b.case_number or "—",
                b.phase.value,
                (b.arbitration_manager or "—")[:25],
            )
        console.print(table)

    # Ошибки
    if report.errors:
        console.print()
        for err in report.errors:
            console.print(f"[yellow]  {err}[/yellow]")

    console.print()


def _style_header(ws, col_count: int):
    """Стилизация заголовков Excel."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Автоширина колонок."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _ru_label(key: str) -> str:
    """Перевод ключей summary на русский."""
    labels = {
        "debtor": "Должник",
        "checked_at": "Дата проверки",
        "enforcements_count": "Исполнительных производств",
        "enforcement_debt_total": "Общий долг по ИП (руб.)",
        "court_cases_count": "Судебных дел",
        "court_claims_total": "Общая сумма исков (руб.)",
        "bankruptcies_count": "Банкротств",
        "has_active_bankruptcy": "Активное банкротство",
        "errors": "Ошибки при проверке",
    }
    return labels.get(key, key)
